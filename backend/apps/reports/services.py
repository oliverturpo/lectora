"""
Servicios para generacion de reportes
"""
import os
from io import BytesIO
from datetime import datetime
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER
from apps.attendance.models import DailySession, Attendance
from apps.attendance.utils import get_system_config
from apps.students.models import Student


def get_institution_name():
    """Obtiene el nombre de la institución desde la configuración del sistema"""
    config = get_system_config()
    return config.get('institution_name', settings.INSTITUTION_CONFIG.get('NAME', 'IES'))

# Rutas de imagenes para reportes
STATIC_DIR = os.path.join(settings.BASE_DIR, 'static', 'images')
LOGO_PATH = os.path.join(STATIC_DIR, 'logo.png')
ESCUDO_PATH = os.path.join(STATIC_DIR, 'escudo.png')


def get_attendance_data(date, grade=None, status=None):
    """Obtiene los datos de asistencia para una fecha con estadísticas calculadas"""
    session = DailySession.objects.filter(date=date).first()
    empty_stats = {'total': 0, 'present': 0, 'late': 0, 'absent': 0, 'percentage': 0}
    if not session:
        return None, [], empty_stats

    attendances = Attendance.objects.filter(session=session).select_related('student')

    if grade:
        attendances = attendances.filter(student__grade=grade)

    if status:
        attendances = attendances.filter(status=status)

    data = []
    # Contadores para estadísticas reales
    stats = {
        'total': 0,
        'present': 0,
        'late': 0,
        'absent': 0
    }

    for att in attendances.order_by('student__grade', 'student__section', 'student__paternal_surname'):
        # Convertir timestamp UTC a hora local de Perú
        hora_local = '-'
        if att.scan_timestamp:
            local_time = timezone.localtime(att.scan_timestamp)
            hora_local = local_time.strftime('%H:%M:%S')

        data.append({
            'dni': att.student.dni,
            'nombre': att.student.full_name,
            'grado': att.student.grade,
            'seccion': att.student.section,
            'estado': att.get_status_display(),
            'estado_key': att.status,  # Clave para conteo
            'hora': hora_local,
        })

        # Contar estadísticas reales
        stats['total'] += 1
        if att.status == 'present':
            stats['present'] += 1
        elif att.status == 'late':
            stats['late'] += 1
        elif att.status == 'absent':
            stats['absent'] += 1

    # Calcular porcentaje de asistencia (presentes + tardanzas = asistieron)
    if stats['total'] > 0:
        attended = stats['present'] + stats['late']
        stats['percentage'] = round((attended / stats['total']) * 100, 1)
    else:
        stats['percentage'] = 0

    return session, data, stats


def generate_excel_report(date, grade=None, status=None):
    """Genera reporte en formato Excel"""
    session, data, stats = get_attendance_data(date, grade, status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titulo
    institution = get_institution_name()
    ws.merge_cells('A1:F1')
    ws['A1'] = institution
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:F2')
    status_names = {'present': 'PRESENTES', 'late': 'TARDANZAS', 'absent': 'FALTAS'}
    title = f"REPORTE DE ASISTENCIA - {date.strftime('%d/%m/%Y')}"
    if status:
        title = f"REPORTE DE {status_names.get(status, status.upper())} - {date.strftime('%d/%m/%Y')}"
    if grade:
        title += f" - {grade} Secundaria"
    ws['A2'] = title
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Estadisticas (calculadas desde los datos reales)
    if stats['total'] > 0:
        ws['A4'] = f"Total: {stats['total']}"
        ws['B4'] = f"Presentes: {stats['present']}"
        ws['C4'] = f"Tardanzas: {stats['late']}"
        ws['D4'] = f"Faltas: {stats['absent']}"
        ws['E4'] = f"Asistencia: {stats['percentage']}%"

    # Headers de la tabla
    headers = ['DNI', 'Nombre Completo', 'Grado', 'Seccion', 'Estado', 'Hora']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    status_fills = {
        'Presente': PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid"),
        'Tardanza': PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid"),
        'Falta': PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid"),
    }

    for row_num, record in enumerate(data, 7):
        ws.cell(row=row_num, column=1, value=record['dni']).border = thin_border
        ws.cell(row=row_num, column=2, value=record['nombre']).border = thin_border
        ws.cell(row=row_num, column=3, value=record['grado']).border = thin_border
        ws.cell(row=row_num, column=4, value=record['seccion']).border = thin_border

        status_cell = ws.cell(row=row_num, column=5, value=record['estado'])
        status_cell.border = thin_border
        status_cell.fill = status_fills.get(record['estado'], PatternFill())

        ws.cell(row=row_num, column=6, value=record['hora']).border = thin_border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_pdf_report(date, grade=None, status=None):
    """Genera reporte en formato PDF profesional con header y footer institucional fijos"""
    session, data, stats = get_attendance_data(date, grade, status)

    buffer = BytesIO()
    page_width, page_height = A4

    # Configurar documento con margenes para header y footer fijos
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=5.5*cm,  # Espacio reservado para header fijo
        bottomMargin=2*cm,  # Espacio para footer
        leftMargin=1.5*cm,
        rightMargin=1.5*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # ============ ESTADISTICAS (calculadas desde datos reales) ============
    if stats['total'] > 0:
        stats_data = [[
            f"Total\n{stats['total']}",
            f"Presentes\n{stats['present']}",
            f"Tardanzas\n{stats['late']}",
            f"Faltas\n{stats['absent']}",
            f"Asistencia\n{stats['percentage']}%"
        ]]
        stats_table = Table(stats_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        stats_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#e0e7ff')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#dcfce7')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#fef3c7')),
            ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#fee2e2')),
            ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#dbeafe')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 15))

    # ============ FILTROS APLICADOS ============
    status_names = {'present': 'PRESENTES', 'late': 'TARDANZAS', 'absent': 'FALTAS'}
    filters_text = ""
    if grade:
        filters_text += f"Grado: {grade}° Secundaria"
    if status:
        if filters_text:
            filters_text += " | "
        filters_text += f"Estado: {status_names.get(status, status.upper())}"

    if filters_text:
        filter_style = ParagraphStyle(
            'FilterStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666'),
            spaceAfter=10
        )
        elements.append(Paragraph(f"Filtros: {filters_text}", filter_style))
        elements.append(Spacer(1, 10))

    # ============ TABLA DE DATOS ============
    table_data = [['N°', 'DNI', 'Nombre Completo', 'Grado', 'Sec.', 'Estado', 'Hora']]
    for idx, record in enumerate(data, 1):
        table_data.append([
            str(idx),
            record['dni'],
            record['nombre'][:35],
            record['grado'],
            record['seccion'],
            record['estado'],
            record['hora']
        ])

    if len(table_data) > 1:
        table = Table(
            table_data,
            colWidths=[1*cm, 2*cm, 6*cm, 1.5*cm, 1.2*cm, 2*cm, 1.8*cm],
            repeatRows=1  # Repetir header de tabla en cada pagina
        )

        # Estilos de la tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Nombre alineado a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ])

        # Colores por estado
        for i, record in enumerate(data, 1):
            if record['estado'] == 'Presente':
                style.add('BACKGROUND', (5, i), (5, i), colors.HexColor('#dcfce7'))
                style.add('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#166534'))
            elif record['estado'] == 'Tardanza':
                style.add('BACKGROUND', (5, i), (5, i), colors.HexColor('#fef3c7'))
                style.add('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#92400e'))
            elif record['estado'] == 'Falta':
                style.add('BACKGROUND', (5, i), (5, i), colors.HexColor('#fee2e2'))
                style.add('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#991b1b'))

        table.setStyle(style)
        elements.append(table)
    else:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666'),
            spaceBefore=30
        )
        elements.append(Paragraph("No hay registros de asistencia para esta fecha.", no_data_style))

    # ============ FUNCION PARA HEADER Y FOOTER FIJOS ============
    institution_name = get_institution_name()

    def add_header_footer(canvas, doc):
        """Dibuja header y footer fijos en cada pagina"""
        canvas.saveState()

        # ========== HEADER FIJO ==========
        img_size = 1.8*cm
        header_top = page_height - 1*cm  # Posicion superior del header

        # Logo izquierdo
        if os.path.exists(LOGO_PATH):
            canvas.drawImage(
                LOGO_PATH,
                1.5*cm,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        # Escudo derecho
        if os.path.exists(ESCUDO_PATH):
            canvas.drawImage(
                ESCUDO_PATH,
                page_width - 1.5*cm - img_size,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        # Textos centrales del header
        center_x = page_width / 2

        # Titulo institucion
        canvas.setFont('Helvetica-Bold', 11)
        canvas.setFillColor(colors.HexColor('#1e3a5f'))
        canvas.drawCentredString(center_x, header_top - 0.5*cm, "INSTITUCIÓN EDUCATIVA SECUNDARIA")
        canvas.drawCentredString(center_x, header_top - 1*cm, institution_name.upper())

        # Subtitulo
        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(colors.black)
        canvas.drawCentredString(center_x, header_top - 1.7*cm, "REGISTRO DE ASISTENCIA")

        # Fecha
        canvas.setFont('Helvetica', 10)
        canvas.drawCentredString(center_x, header_top - 2.2*cm, f"FECHA: {date.strftime('%d / %m / %Y')}")

        # Linea separadora del header
        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1.5)
        canvas.line(1.5*cm, header_top - 2.8*cm, page_width - 1.5*cm, header_top - 2.8*cm)

        # ========== FOOTER FIJO ==========
        # Linea del footer
        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1)
        canvas.line(1.5*cm, 1.5*cm, page_width - 1.5*cm, 1.5*cm)

        # Texto del footer
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))

        # Izquierda: "Hecho en Django + React"
        canvas.drawString(1.5*cm, 1*cm, "Hecho en Django + React")

        # Centro: Fecha de generacion
        gen_text = f"Generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        canvas.drawCentredString(page_width / 2, 1*cm, gen_text)

        # Derecha: Numero de pagina
        page_num = f"Página {doc.page}"
        canvas.drawRightString(page_width - 1.5*cm, 1*cm, page_num)

        canvas.restoreState()

    doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    buffer.seek(0)

    return buffer


def generate_nomina_by_grade_pdf(grade, section=None):
    """
    Genera PDF con nómina de estudiantes por grado y sección
    """
    buffer = BytesIO()
    page_width, page_height = A4

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=5*cm,
        bottomMargin=2*cm,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Filtrar estudiantes
    students = Student.objects.filter(is_active=True, grade=grade)
    if section:
        students = students.filter(section=section)
    students = students.order_by('section', 'paternal_surname', 'maternal_surname')

    # Tabla: N°, DNI, Apellidos y Nombres, Grado, Sección
    table_data = [['N°', 'DNI', 'Apellidos y Nombres', 'Grado', 'Sección']]
    for idx, student in enumerate(students, 1):
        table_data.append([
            str(idx),
            student.dni,
            student.full_name,
            student.grade,
            student.section
        ])

    if len(table_data) > 1:
        table = Table(
            table_data,
            colWidths=[1*cm, 2.5*cm, 8*cm, 1.5*cm, 1.5*cm],
            repeatRows=1
        )

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ])

        table.setStyle(style)
        elements.append(table)
    else:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666'),
            spaceBefore=30
        )
        elements.append(Paragraph("No hay estudiantes registrados para este grado/sección.", no_data_style))

    # Footer con total
    elements.append(Spacer(1, 20))
    total_style = ParagraphStyle(
        'Total',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph(f"<b>Total de estudiantes: {len(table_data) - 1}</b>", total_style))

    institution_name = get_institution_name()

    def add_nomina_header_footer(canvas, doc):
        """Dibuja header y footer para nómina por grado"""
        canvas.saveState()

        # Header
        img_size = 1.8*cm
        header_top = page_height - 1*cm

        if os.path.exists(LOGO_PATH):
            canvas.drawImage(
                LOGO_PATH,
                1.5*cm,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        if os.path.exists(ESCUDO_PATH):
            canvas.drawImage(
                ESCUDO_PATH,
                page_width - 1.5*cm - img_size,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        center_x = page_width / 2
        canvas.setFont('Helvetica-Bold', 11)
        canvas.setFillColor(colors.HexColor('#1e3a5f'))
        canvas.drawCentredString(center_x, header_top - 0.5*cm, "INSTITUCIÓN EDUCATIVA SECUNDARIA")
        canvas.drawCentredString(center_x, header_top - 1*cm, institution_name.upper())

        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(colors.black)
        if section:
            canvas.drawCentredString(center_x, header_top - 1.7*cm, f"NÓMINA DE ESTUDIANTES - {grade}° Secundaria - Sección {section}")
        else:
            canvas.drawCentredString(center_x, header_top - 1.7*cm, f"NÓMINA DE ESTUDIANTES - {grade}° Secundaria")

        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1.5)
        canvas.line(1.5*cm, header_top - 2.3*cm, page_width - 1.5*cm, header_top - 2.3*cm)

        # Footer
        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1)
        canvas.line(1.5*cm, 1.5*cm, page_width - 1.5*cm, 1.5*cm)

        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(1.5*cm, 1*cm, "Sistema de Asistencia - Django + React")

        gen_text = f"Generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        canvas.drawCentredString(page_width / 2, 1*cm, gen_text)

        page_num = f"Página {doc.page}"
        canvas.drawRightString(page_width - 1.5*cm, 1*cm, page_num)

        canvas.restoreState()

    doc.build(elements, onFirstPage=add_nomina_header_footer, onLaterPages=add_nomina_header_footer)
    buffer.seek(0)

    return buffer


def generate_nomina_oficial_pdf():
    """
    Genera PDF con nómina oficial de todos los estudiantes del colegio
    """
    buffer = BytesIO()
    page_width, page_height = A4

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=6*cm,
        bottomMargin=2*cm,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Obtener TODOS los estudiantes activos
    students = Student.objects.filter(is_active=True).order_by(
        'grade', 'section', 'paternal_surname', 'maternal_surname'
    )

    # Resumen por grados
    stats_by_grade = {}
    for student in students:
        grade = student.grade
        if grade not in stats_by_grade:
            stats_by_grade[grade] = 0
        stats_by_grade[grade] += 1

    # Tabla de resumen
    summary_data = [['Grado', 'Total Estudiantes']]
    for grade in sorted(stats_by_grade.keys()):
        summary_data.append([f"{grade}° Secundaria", str(stats_by_grade[grade])])

    summary_table = Table(summary_data, colWidths=[8*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Tabla principal: N°, DNI, Apellidos y Nombres, Grado, Sección
    table_data = [['N°', 'DNI', 'Apellidos y Nombres', 'Grado', 'Sección']]
    for idx, student in enumerate(students, 1):
        table_data.append([
            str(idx),
            student.dni,
            student.full_name,
            student.grade,
            student.section
        ])

    if len(table_data) > 1:
        table = Table(
            table_data,
            colWidths=[1*cm, 2.5*cm, 8*cm, 1.5*cm, 1.5*cm],
            repeatRows=1
        )

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ])

        table.setStyle(style)
        elements.append(table)
    else:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666'),
            spaceBefore=30
        )
        elements.append(Paragraph("No hay estudiantes registrados.", no_data_style))

    # Footer con total
    elements.append(Spacer(1, 20))
    total_style = ParagraphStyle(
        'Total',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph(f"<b>Total de estudiantes: {len(table_data) - 1}</b>", total_style))

    institution_name = get_institution_name()

    def add_oficial_header_footer(canvas, doc):
        """Dibuja header y footer para nómina oficial"""
        canvas.saveState()

        # Header
        img_size = 1.8*cm
        header_top = page_height - 1*cm

        if os.path.exists(LOGO_PATH):
            canvas.drawImage(
                LOGO_PATH,
                1.5*cm,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        if os.path.exists(ESCUDO_PATH):
            canvas.drawImage(
                ESCUDO_PATH,
                page_width - 1.5*cm - img_size,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        center_x = page_width / 2
        canvas.setFont('Helvetica-Bold', 11)
        canvas.setFillColor(colors.HexColor('#1e3a5f'))
        canvas.drawCentredString(center_x, header_top - 0.5*cm, "INSTITUCIÓN EDUCATIVA SECUNDARIA")
        canvas.drawCentredString(center_x, header_top - 1*cm, institution_name.upper())

        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(colors.black)
        canvas.drawCentredString(center_x, header_top - 1.7*cm, "NÓMINA OFICIAL DE ESTUDIANTES")

        canvas.setFont('Helvetica', 9)
        year = datetime.now().year
        canvas.drawCentredString(center_x, header_top - 2.2*cm, f"AÑO ACADÉMICO {year}")

        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1.5)
        canvas.line(1.5*cm, header_top - 2.8*cm, page_width - 1.5*cm, header_top - 2.8*cm)

        # Footer
        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1)
        canvas.line(1.5*cm, 1.5*cm, page_width - 1.5*cm, 1.5*cm)

        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(1.5*cm, 1*cm, "Sistema de Asistencia - Django + React")

        gen_text = f"Generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        canvas.drawCentredString(page_width / 2, 1*cm, gen_text)

        page_num = f"Página {doc.page}"
        canvas.drawRightString(page_width - 1.5*cm, 1*cm, page_num)

        canvas.restoreState()

    doc.build(elements, onFirstPage=add_oficial_header_footer, onLaterPages=add_oficial_header_footer)
    buffer.seek(0)

    return buffer


def generate_student_attendance_pdf(student_id):
    """
    Genera PDF con el historial de asistencia de un estudiante
    """
    from apps.students.models import Student

    try:
        student = Student.objects.get(id=student_id, is_active=True)
    except Student.DoesNotExist:
        return None

    buffer = BytesIO()
    page_width, page_height = A4

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=6*cm,
        bottomMargin=2*cm,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Obtener historial de asistencias
    attendances = Attendance.objects.filter(student=student).select_related('session').order_by('-session__date')

    # Calcular estadísticas
    total_days = attendances.count()
    present = attendances.filter(status='present').count()
    late = attendances.filter(status='late').count()
    absent = attendances.filter(status='absent').count()
    attendance_percentage = round((present + late) / total_days * 100, 1) if total_days > 0 else 0

    # Resumen estadístico
    summary_data = [[
        f"Total Días\n{total_days}",
        f"Presentes\n{present}",
        f"Tardanzas\n{late}",
        f"Faltas\n{absent}",
        f"Asistencia\n{attendance_percentage}%"
    ]]
    summary_table = Table(summary_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#dbeafe')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#dcfce7')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#fef3c7')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#fee2e2')),
        ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#e0e7ff')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Tabla de historial
    if total_days > 0:
        table_data = [['N°', 'Fecha', 'Estado', 'Hora', 'Método']]
        for idx, att in enumerate(attendances, 1):
            # Convertir timestamp a hora local
            hora_local = '-'
            if att.scan_timestamp:
                local_time = timezone.localtime(att.scan_timestamp)
                hora_local = local_time.strftime('%H:%M:%S')

            fecha_str = att.session.date.strftime('%d/%m/%Y')

            table_data.append([
                str(idx),
                fecha_str,
                att.get_status_display(),
                hora_local,
                att.get_registration_method_display()
            ])

        table = Table(
            table_data,
            colWidths=[1*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm],
            repeatRows=1
        )

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ])

        # Colores por estado
        for i, att in enumerate(attendances, 1):
            if att.status == 'present':
                style.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#dcfce7'))
                style.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#166534'))
            elif att.status == 'late':
                style.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#fef3c7'))
                style.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#92400e'))
            elif att.status == 'absent':
                style.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#fee2e2'))
                style.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#991b1b'))

        table.setStyle(style)
        elements.append(table)
    else:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666'),
            spaceBefore=30
        )
        elements.append(Paragraph("No hay registros de asistencia para este estudiante.", no_data_style))

    institution_name = get_institution_name()

    def add_student_header_footer(canvas, doc):
        """Dibuja header y footer para reporte de estudiante"""
        canvas.saveState()

        # Header
        img_size = 1.8*cm
        header_top = page_height - 1*cm

        if os.path.exists(LOGO_PATH):
            canvas.drawImage(
                LOGO_PATH,
                1.5*cm,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        if os.path.exists(ESCUDO_PATH):
            canvas.drawImage(
                ESCUDO_PATH,
                page_width - 1.5*cm - img_size,
                header_top - img_size,
                width=img_size,
                height=img_size,
                preserveAspectRatio=True,
                mask='auto'
            )

        center_x = page_width / 2
        canvas.setFont('Helvetica-Bold', 11)
        canvas.setFillColor(colors.HexColor('#1e3a5f'))
        canvas.drawCentredString(center_x, header_top - 0.5*cm, "INSTITUCIÓN EDUCATIVA SECUNDARIA")
        canvas.drawCentredString(center_x, header_top - 1*cm, institution_name.upper())

        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(colors.black)
        canvas.drawCentredString(center_x, header_top - 1.7*cm, "REPORTE DE ASISTENCIA - ESTUDIANTE")

        # Info del estudiante
        canvas.setFont('Helvetica', 9)
        canvas.drawCentredString(center_x, header_top - 2.3*cm, f"Estudiante: {student.full_name}")
        canvas.drawCentredString(center_x, header_top - 2.8*cm, f"DNI: {student.dni} | Grado: {student.grade}° - Sección: {student.section}")

        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1.5)
        canvas.line(1.5*cm, header_top - 3.3*cm, page_width - 1.5*cm, header_top - 3.3*cm)

        # Footer
        canvas.setStrokeColor(colors.HexColor('#1e40af'))
        canvas.setLineWidth(1)
        canvas.line(1.5*cm, 1.5*cm, page_width - 1.5*cm, 1.5*cm)

        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(1.5*cm, 1*cm, "Sistema de Asistencia - Django + React")

        gen_text = f"Generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        canvas.drawCentredString(page_width / 2, 1*cm, gen_text)

        page_num = f"Página {doc.page}"
        canvas.drawRightString(page_width - 1.5*cm, 1*cm, page_num)

        canvas.restoreState()

    doc.build(elements, onFirstPage=add_student_header_footer, onLaterPages=add_student_header_footer)
    buffer.seek(0)

    return buffer


def generate_nomina_oficial_excel():
    """
    Genera Excel con nómina oficial de todos los estudiantes.
    Cada hoja corresponde a un grado y sección existente.
    """
    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)

    # Obtener todos los estudiantes activos ordenados
    students = Student.objects.filter(is_active=True).order_by(
        'grade', 'section', 'paternal_surname', 'maternal_surname'
    )

    # Agrupar por grado y sección
    grades_sections = {}
    for student in students:
        key = f"{student.grade}_{student.section}"
        if key not in grades_sections:
            grades_sections[key] = []
        grades_sections[key].append(student)

    # Crear una hoja por cada grado-sección
    for key in sorted(grades_sections.keys()):
        grade, section = key.split('_')
        sheet_name = f"{grade}° - {section}"
        ws = wb.create_sheet(title=sheet_name)

        students_list = grades_sections[key]
        institution = get_institution_name()

        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = institution
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:E2')
        ws['A2'] = f"NÓMINA DE ESTUDIANTES - {grade}° Secundaria - Sección {section}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A3:E3')
        ws['A3'] = f"Año Académico {datetime.now().year}"
        ws['A3'].alignment = Alignment(horizontal="center")

        # Headers de la tabla
        headers = ['N°', 'DNI', 'Apellidos y Nombres', 'Grado', 'Sección']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos de estudiantes
        for idx, student in enumerate(students_list, 1):
            row = idx + 5
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=2, value=student.dni).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=3, value=student.full_name).border = thin_border

            ws.cell(row=row, column=4, value=student.grade).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=5, value=student.section).border = thin_border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        # Total al final
        total_row = len(students_list) + 6
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws[f'A{total_row}'] = f"Total de estudiantes: {len(students_list)}"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'A{total_row}'].alignment = Alignment(horizontal="center")

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10

    # Crear hoja de resumen al inicio
    summary_ws = wb.create_sheet(title="Resumen", index=0)
    institution = get_institution_name()

    summary_ws.merge_cells('A1:C1')
    summary_ws['A1'] = institution
    summary_ws['A1'].font = title_font
    summary_ws['A1'].alignment = Alignment(horizontal="center")

    summary_ws.merge_cells('A2:C2')
    summary_ws['A2'] = "RESUMEN DE NÓMINA OFICIAL"
    summary_ws['A2'].font = subtitle_font
    summary_ws['A2'].alignment = Alignment(horizontal="center")

    summary_ws.merge_cells('A3:C3')
    summary_ws['A3'] = f"Año Académico {datetime.now().year}"
    summary_ws['A3'].alignment = Alignment(horizontal="center")

    # Headers del resumen
    summary_headers = ['Grado', 'Sección', 'Total Estudiantes']
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos del resumen
    row_num = 6
    total_general = 0
    for key in sorted(grades_sections.keys()):
        grade, section = key.split('_')
        count = len(grades_sections[key])
        total_general += count

        summary_ws.cell(row=row_num, column=1, value=f"{grade}°").border = thin_border
        summary_ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")

        summary_ws.cell(row=row_num, column=2, value=section).border = thin_border
        summary_ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center")

        summary_ws.cell(row=row_num, column=3, value=count).border = thin_border
        summary_ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")

        row_num += 1

    # Total general
    summary_ws.merge_cells(f'A{row_num}:B{row_num}')
    summary_ws[f'A{row_num}'] = "TOTAL GENERAL"
    summary_ws[f'A{row_num}'].font = Font(bold=True)
    summary_ws[f'A{row_num}'].alignment = Alignment(horizontal="center")
    summary_ws[f'A{row_num}'].border = thin_border
    summary_ws[f'B{row_num}'].border = thin_border

    summary_ws.cell(row=row_num, column=3, value=total_general).border = thin_border
    summary_ws.cell(row=row_num, column=3).font = Font(bold=True)
    summary_ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")

    # Ajustar anchos
    summary_ws.column_dimensions['A'].width = 15
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 20

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_complete_attendance_excel(grade_filter=None, section_filter=None):
    """
    Genera Excel con asistencia completa de todos los estudiantes.
    Cada hoja corresponde a un grado y sección.
    Columnas: DNI | Nombre | Grado | Sección | Fecha1 | Fecha2 | ... | % Asistencia
    Valores: P (presente), T (tardanza), F (falta)
    Porcentaje: Solo cuenta las "P" dividido entre total de días
    """
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)

    # Colores para estados
    present_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    late_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    absent_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    percent_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")

    # Obtener todos los estudiantes activos
    students_query = Student.objects.filter(is_active=True)
    if grade_filter:
        students_query = students_query.filter(grade=grade_filter)
    if section_filter:
        students_query = students_query.filter(section=section_filter)

    students = students_query.order_by('grade', 'section', 'paternal_surname', 'maternal_surname')

    # Obtener todas las sesiones ordenadas por fecha
    sessions = DailySession.objects.all().order_by('date')
    session_dates = list(sessions.values_list('id', 'date'))

    # Agrupar estudiantes por grado y sección
    grades_sections = {}
    for student in students:
        key = f"{student.grade}_{student.section}"
        if key not in grades_sections:
            grades_sections[key] = []
        grades_sections[key].append(student)

    # Crear diccionario de asistencias para acceso rápido
    # {student_id: {session_id: status}}
    attendance_dict = {}
    all_attendances = Attendance.objects.select_related('student', 'session').all()
    for att in all_attendances:
        if att.student_id not in attendance_dict:
            attendance_dict[att.student_id] = {}
        attendance_dict[att.student_id][att.session_id] = att.status

    institution = get_institution_name()

    # Crear una hoja por cada grado-sección
    for key in sorted(grades_sections.keys()):
        grade, section = key.split('_')
        sheet_name = f"{grade} - {section}"
        ws = wb.create_sheet(title=sheet_name)

        students_list = grades_sections[key]

        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = institution
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:E2')
        ws['A2'] = f"REGISTRO DE ASISTENCIA - {grade} Secundaria - Sección {section}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A3:E3')
        ws['A3'] = f"Año Académico {datetime.now().year}"
        ws['A3'].alignment = Alignment(horizontal="center")

        # Headers fijos
        headers = ['N°', 'DNI', 'Apellidos y Nombres', 'Grado', 'Sección']

        # Agregar columnas de fechas
        for session_id, session_date in session_dates:
            headers.append(session_date.strftime('%d/%m'))

        # Agregar columna de porcentaje
        headers.append('% Asist.')

        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos de estudiantes
        for idx, student in enumerate(students_list, 1):
            row = idx + 5

            # N°
            cell = ws.cell(row=row, column=1, value=idx)
            cell.border = thin_border
            cell.alignment = center_alignment

            # DNI
            cell = ws.cell(row=row, column=2, value=student.dni)
            cell.border = thin_border
            cell.alignment = center_alignment

            # Nombre completo
            cell = ws.cell(row=row, column=3, value=student.full_name)
            cell.border = thin_border

            # Grado
            cell = ws.cell(row=row, column=4, value=student.grade)
            cell.border = thin_border
            cell.alignment = center_alignment

            # Sección
            cell = ws.cell(row=row, column=5, value=student.section)
            cell.border = thin_border
            cell.alignment = center_alignment

            # Asistencias por fecha
            total_days = 0
            present_count = 0
            col_offset = 6

            student_attendance = attendance_dict.get(student.id, {})

            for session_id, session_date in session_dates:
                status = student_attendance.get(session_id)
                cell = ws.cell(row=row, column=col_offset)
                cell.border = thin_border
                cell.alignment = center_alignment

                if status:
                    total_days += 1
                    if status == 'present':
                        cell.value = 'P'
                        cell.fill = present_fill
                        present_count += 1
                    elif status == 'late':
                        cell.value = 'T'
                        cell.fill = late_fill
                    elif status == 'absent':
                        cell.value = 'F'
                        cell.fill = absent_fill
                else:
                    cell.value = '-'

                col_offset += 1

            # Calcular porcentaje (solo P cuenta)
            percentage = 0
            if total_days > 0:
                percentage = round((present_count / total_days) * 100)

            cell = ws.cell(row=row, column=col_offset, value=f"{percentage}%")
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = percent_fill
            cell.font = Font(bold=True)

        # Total al final
        total_row = len(students_list) + 6
        ws.cell(row=total_row, column=1, value=f"Total: {len(students_list)} estudiantes")
        ws.cell(row=total_row, column=1).font = Font(bold=True)

        # Leyenda
        legend_row = total_row + 2
        ws.cell(row=legend_row, column=1, value="Leyenda:")
        ws.cell(row=legend_row, column=1).font = Font(bold=True)

        ws.cell(row=legend_row + 1, column=1, value="P = Presente")
        ws.cell(row=legend_row + 1, column=1).fill = present_fill

        ws.cell(row=legend_row + 1, column=2, value="T = Tardanza")
        ws.cell(row=legend_row + 1, column=2).fill = late_fill

        ws.cell(row=legend_row + 1, column=3, value="F = Falta")
        ws.cell(row=legend_row + 1, column=3).fill = absent_fill

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8

        # Ancho para columnas de fechas y porcentaje
        for i in range(6, 6 + len(session_dates) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 7

    # Si no hay datos, crear hoja vacía con mensaje
    if not grades_sections:
        ws = wb.create_sheet(title="Sin datos")
        ws['A1'] = "No hay estudiantes registrados con los filtros seleccionados."

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
